import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import os

def conectar_banco():
    try:
        engine = create_engine('mysql+mysqlconnector://root:12344321@localhost/venda_db')
        print("Conectado ao MySQL com sucesso!")
        return engine
    except Exception as e:
        print(f"Erro ao conectar ao MySQL: {e}")
        return None

def extrair_dados(engine):
    query = "SELECT * FROM vendas;"
    df = pd.read_sql(query, engine)
    return df

def gerar_relatorio_excel(df_vendas, vendas_por_dia, vendas_por_produto):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Resumo Vendas"

    for r in dataframe_to_rows(df_vendas, index=False, header=True):
        ws1.append(r)

    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    ws2 = wb.create_sheet(title="Vendas por Dia")
    for r in dataframe_to_rows(vendas_por_dia, index=False, header=True):
        ws2.append(r)

    chart = BarChart()
    chart.title = "Total de Vendas por Dia"
    chart.x_axis.title = "Data"
    chart.y_axis.title = "Total Vendas"
    data = Reference(ws2, min_col=2, min_row=2, max_row=ws2.max_row, max_col=2)
    labels = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    ws2.add_chart(chart, "E5")

    ws3 = wb.create_sheet(title="Vendas por Produto")
    for r in dataframe_to_rows(vendas_por_produto, index=False, header=True):
        ws3.append(r)

    nome_arquivo = "relatorio_vendas.xlsx"
    wb.save(nome_arquivo)
    print(f"\nRelatório Excel gerado com sucesso: {nome_arquivo}")
    return nome_arquivo

def enviar_email(arquivo_excel):
    remetente = 'rodzmaciel21@gmail.com'
    senha = 'zvrz fvay vqsv wjds'
    destinatario = 'rodzmaciel21@gmail.com'
    assunto = 'Relatório de Vendas Diário'
    mensagem = 'Segue em anexo o relatório de vendas atualizado.'

    msg = MIMEMultipart()
    msg['From'] = remetente
    msg['To'] = destinatario
    msg['Subject'] = assunto
    msg.attach(MIMEText(mensagem, 'plain'))

    with open(arquivo_excel, 'rb') as f:
        parte = MIMEApplication(f.read(), Name=os.path.basename(arquivo_excel))
        parte['Content-Disposition'] = f'attachment; filename="{os.path.basename(arquivo_excel)}"'
        msg.attach(parte)

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as servidor:
            servidor.starttls()
            servidor.login(remetente, senha)
            servidor.send_message(msg)
            print("E-mail enviado com sucesso!")
    except Exception as e:
        print(f"Erro ao enviar o e-mail: {e}")

if __name__ == "__main__":
    engine = conectar_banco()
    if engine:
        df_vendas = extrair_dados(engine)

        df_vendas['data'] = pd.to_datetime(df_vendas['data'], errors='coerce')
        df_vendas['quantidade'] = pd.to_numeric(df_vendas['quantidade'], errors='coerce')
        df_vendas['preco'] = pd.to_numeric(df_vendas['preco'], errors='coerce')
        df_vendas.fillna(0, inplace=True)
        df_vendas['total_venda'] = df_vendas['quantidade'] * df_vendas['preco']

        vendas_por_dia = df_vendas.groupby('data')['total_venda'].sum().reset_index()
        vendas_por_produto = df_vendas.groupby('produto')['total_venda'].sum().reset_index()

        nome_arquivo = gerar_relatorio_excel(df_vendas, vendas_por_dia, vendas_por_produto)

        enviar_email(nome_arquivo)

        engine.dispose()
